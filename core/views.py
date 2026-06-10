from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.utils import timezone
from django.db.models import Q
import io

from .models import UserProfile, Product, Order
from .forms import SignupForm, EmployeeCreationForm, ProductForm, OrderForm, AdminCreationForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ReportLab Imports for Receipt Generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def home_view(request):
    products = Product.objects.all().order_by('-id')
    signup_form = SignupForm()
    
    search_q = request.GET.get('search_products', '').strip()
    if search_q:
        products = products.filter(Q(name__icontains=search_q) | Q(description__icontains=search_q))
        
    # Check if there is an error in login/signup from redirects
    return render(request, 'core/home.html', {
        'products': products,
        'signup_form': signup_form,
        'search_products_query': search_q,
    })

def signup_view(request):
    if request.method == 'POST':
        form = SignupForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f"Welcome {user.username}! Registration successful.")
            return redirect('dashboard')
        else:
            products = Product.objects.all().order_by('-id')
            return render(request, 'core/home.html', {
                'products': products,
                'signup_form': form,
                'open_signup_modal': True
            })
    return redirect('home')

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
            products = Product.objects.all().order_by('-id')
            return render(request, 'core/home.html', {
                'products': products,
                'signup_form': SignupForm(),
                'open_login_modal': True
            })
    return redirect('home')

def logout_view(request):
    logout(request)
    messages.info(request, "Successfully logged out.")
    return redirect('home')

@login_required
def dashboard_view(request):
    user = request.user
    role = user.role

    if user.role == 'admin' or user.is_superuser:
        vendors = UserProfile.objects.filter(role='vendor').order_by('-id')
        employees = UserProfile.objects.filter(role__in=['shipper', 'out_for_delivery', 'delivered']).order_by('-id')
        
        # Vendor Search
        search_vendors = request.GET.get('search_vendors', '').strip()
        if search_vendors:
            vendors = vendors.filter(Q(username__icontains=search_vendors) | Q(email__icontains=search_vendors))
            
        # Logistics Search (Search by username, email, or role)
        search_employees = request.GET.get('search_employees', '').strip()
        if search_employees:
            employees = employees.filter(
                Q(username__icontains=search_employees) |
                Q(email__icontains=search_employees) |
                Q(role__icontains=search_employees)
            )
            
        # Vendor filter for logistics employees
        filter_vendor_id = request.GET.get('filter_vendor', '').strip()
        if filter_vendor_id:
            employees = employees.filter(created_by_id=filter_vendor_id)
            
        # Get list of all vendors for the dropdown filter
        all_vendors = UserProfile.objects.filter(role='vendor').order_by('username')
        
        admin_form = AdminCreationForm()
        
        return render(request, 'core/admin_dashboard.html', {
            'vendors': vendors,
            'employees': employees,
            'all_vendors': all_vendors,
            'search_vendors': search_vendors,
            'search_employees': search_employees,
            'filter_vendor': filter_vendor_id,
            'admin_form': admin_form,
            'active_tab': request.GET.get('tab', 'logistic_accounts')
        })

    waiting_approval = False
    if role == 'vendor' and not user.is_approved:
        waiting_approval = True
    elif role in ['shipper', 'out_for_delivery', 'delivered']:
        if not user.is_approved or (user.created_by and not user.created_by.is_approved):
            waiting_approval = True

    if role == 'buyer':
        orders = Order.objects.filter(buyer=user).order_by('-order_confirmed_at')
        products = Product.objects.all().order_by('-id')
        
        search_products = request.GET.get('search_products', '').strip()
        if search_products:
            products = products.filter(Q(name__icontains=search_products) | Q(description__icontains=search_products))
            
        search_orders = request.GET.get('search_orders', '').strip()
        if search_orders:
            orders = orders.filter(
                Q(product__name__icontains=search_orders) |
                Q(id__icontains=search_orders) |
                Q(status__icontains=search_orders)
            )
            
        return render(request, 'core/buyer_dashboard.html', {
            'orders': orders,
            'products': products,
            'search_products': search_products,
            'search_orders': search_orders,
            'active_tab': request.GET.get('tab', 'products')
        })

    elif role == 'vendor':
        # Vendor's list of created employees is filtered by vendor
        products = Product.objects.filter(vendor=user).order_by('-id')
        orders = Order.objects.filter(product__vendor=user).order_by('-order_confirmed_at')
        employees = UserProfile.objects.filter(created_by=user).order_by('-id')
        
        search_products = request.GET.get('search_products', '').strip()
        if search_products:
            products = products.filter(Q(name__icontains=search_products) | Q(description__icontains=search_products))
            
        search_orders = request.GET.get('search_orders', '').strip()
        if search_orders:
            orders = orders.filter(
                Q(product__name__icontains=search_orders) |
                Q(id__icontains=search_orders) |
                Q(status__icontains=search_orders) |
                Q(buyer__username__icontains=search_orders)
            )
            
        search_employees = request.GET.get('search_employees', '').strip()
        if search_employees:
            employees = employees.filter(
                Q(username__icontains=search_employees) |
                Q(email__icontains=search_employees) |
                Q(role__icontains=search_employees)
            )
        
        # Forms
        product_form = ProductForm()
        employee_form = EmployeeCreationForm()
        
        return render(request, 'core/vendor_dashboard.html', {
            'products': products,
            'orders': orders,
            'employees': employees,
            'product_form': product_form,
            'employee_form': employee_form,
            'search_products': search_products,
            'search_orders': search_orders,
            'search_employees': search_employees,
            'waiting_approval': waiting_approval,
            'active_tab': request.GET.get('tab', 'products')
        })

    elif role == 'shipper':
        orders = Order.objects.filter(status='confirmed').order_by('order_confirmed_at')
        all_orders = Order.objects.filter(status='shipped').order_by('-shipped_at')
        return render(request, 'core/shipper_dashboard.html', {
            'orders': orders,
            'all_orders': all_orders,
            'waiting_approval': waiting_approval
        })

    elif role == 'out_for_delivery':
        orders = Order.objects.filter(status='shipped').order_by('shipped_at')
        all_orders = Order.objects.filter(status='out_for_delivery').order_by('-out_for_delivery_at')
        return render(request, 'core/out_for_delivery_dashboard.html', {
            'orders': orders,
            'all_orders': all_orders,
            'waiting_approval': waiting_approval
        })

    elif role == 'delivered':
        orders = Order.objects.filter(status='out_for_delivery').order_by('out_for_delivery_at')
        all_orders = Order.objects.filter(status='delivered').order_by('-delivered_at')
        return render(request, 'core/delivered_dashboard.html', {
            'orders': orders,
            'all_orders': all_orders,
            'waiting_approval': waiting_approval
        })

    else:
        # Fallback to Admin or others
        return redirect('home')

@login_required
def vendor_add_product(request):
    if request.user.role != 'vendor':
        messages.error(request, "Permission denied.")
        return redirect('dashboard')
    if not request.user.is_approved:
        messages.error(request, "Your account is waiting for approval.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.vendor = request.user
            product.save()
            messages.success(request, f"Product '{product.name}' added successfully!")
            return redirect('/dashboard/?tab=products')
        else:
            messages.error(request, "Error adding product. Please verify fields.")
            products = Product.objects.filter(vendor=request.user).order_by('-id')
            orders = Order.objects.filter(product__vendor=request.user).order_by('-order_confirmed_at')
            employees = UserProfile.objects.filter(created_by=request.user).order_by('-id')
            employee_form = EmployeeCreationForm()
            return render(request, 'core/vendor_dashboard.html', {
                'products': products,
                'orders': orders,
                'employees': employees,
                'product_form': form,
                'employee_form': employee_form,
                'waiting_approval': not request.user.is_approved,
                'active_tab': 'add_product'
            })
    return redirect('dashboard')

@login_required
def vendor_create_employee(request):
    if request.user.role != 'vendor':
        messages.error(request, "Permission denied.")
        return redirect('dashboard')
    if not request.user.is_approved:
        messages.error(request, "Your account is waiting for approval.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = EmployeeCreationForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.created_by = request.user
            employee.save()
            messages.success(request, f"Logistics Employee ID for '{employee.username}' (Role: {employee.get_role_display()}) created successfully!")
            return redirect('/dashboard/?tab=employees')
        else:
            messages.error(request, "Error creating employee. Please check details.")
            products = Product.objects.filter(vendor=request.user).order_by('-id')
            orders = Order.objects.filter(product__vendor=request.user).order_by('-order_confirmed_at')
            employees = UserProfile.objects.filter(created_by=request.user).order_by('-id')
            product_form = ProductForm()
            return render(request, 'core/vendor_dashboard.html', {
                'products': products,
                'orders': orders,
                'employees': employees,
                'product_form': product_form,
                'employee_form': form,
                'waiting_approval': not request.user.is_approved,
                'active_tab': 'employees'
            })
    return redirect('dashboard')

@login_required
def checkout_view(request, product_id):
    if request.user.role != 'buyer':
        messages.error(request, "Only Buyers can purchase products.")
        return redirect('dashboard')

    product = get_object_or_404(Product, id=product_id)
    if product.available_quantity <= 0:
        messages.error(request, "This product is currently out of stock.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            payment_method = form.cleaned_data['payment_method']
            
            if payment_method != 'cod':
                messages.error(request, "Only Cash on Delivery is currently supported.")
                return render(request, 'core/checkout.html', {'product': product, 'form': form})

            if quantity > product.available_quantity:
                messages.error(request, f"Only {product.available_quantity} items are available in stock.")
                return render(request, 'core/checkout.html', {'product': product, 'form': form})

            # Create Order
            order = Order(
                product=product,
                buyer=request.user,
                quantity=quantity,
                total_price=product.discounted_price * quantity,
                payment_method='Cash on Delivery',
                status='confirmed',
                order_confirmed_at=timezone.now()
            )
            order.save()

            # Deduct quantity
            product.available_quantity -= quantity
            product.save()

            messages.success(request, f"Order placed successfully! Order ID: #{order.id}")
            return redirect('/dashboard/?tab=orders')
    else:
        form = OrderForm()

    return render(request, 'core/checkout.html', {'product': product, 'form': form})

@login_required
def update_status_shipper(request, order_id):
    if request.user.role != 'shipper':
        messages.error(request, "Permission denied.")
        return redirect('dashboard')
    if request.user.created_by and not request.user.created_by.is_approved:
        messages.error(request, "Operations suspended. Creator Vendor is waiting for approval / disapproved.")
        return redirect('dashboard')
    
    order = get_object_or_404(Order, id=order_id, status='confirmed')
    order.status = 'shipped'
    order.shipped_at = timezone.now()
    order.save()
    messages.success(request, f"Order #{order.id} marked as Shipped successfully!")
    return redirect('dashboard')

@login_required
def update_status_out_for_delivery(request, order_id):
    if request.user.role != 'out_for_delivery':
        messages.error(request, "Permission denied.")
        return redirect('dashboard')
    if request.user.created_by and not request.user.created_by.is_approved:
        messages.error(request, "Operations suspended. Creator Vendor is waiting for approval / disapproved.")
        return redirect('dashboard')
    
    order = get_object_or_404(Order, id=order_id, status='shipped')
    order.status = 'out_for_delivery'
    order.out_for_delivery_at = timezone.now()
    order.save()
    messages.success(request, f"Order #{order.id} marked as Out for Delivery / On the Way!")
    return redirect('dashboard')

@login_required
def verify_otp_and_deliver(request, order_id):
    if request.user.role != 'delivered':
        messages.error(request, "Permission denied.")
        return redirect('dashboard')
    if request.user.created_by and not request.user.created_by.is_approved:
        messages.error(request, "Operations suspended. Creator Vendor is waiting for approval / disapproved.")
        return redirect('dashboard')
    
    order = get_object_or_404(Order, id=order_id, status='out_for_delivery')
    
    if request.method == 'POST':
        input_otp = request.POST.get('otp')
        if input_otp == order.otp:
            order.status = 'delivered'
            order.delivered_at = timezone.now()
            order.save()
            messages.success(request, f"OTP Verified! Order #{order.id} marked as Delivered successfully.")
        else:
            messages.error(request, f"Invalid OTP for Order #{order.id}. Delivery failed.")
            
    return redirect('dashboard')

@login_required
def generate_receipt_pdf(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    user = request.user
    
    # Authorization: buyer, vendor of product, logistics roles, or admin
    is_authorized = (
        user == order.buyer or 
        user == order.product.vendor or 
        user.role in ['shipper', 'out_for_delivery', 'delivered'] or
        user.is_superuser
    )
    if not is_authorized:
        raise Http404("You are not authorized to view this receipt.")

    # Generate PDF Response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Receipt_Order_{order.id}.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor("#1A365D"),
        spaceAfter=15,
        alignment=1 # Center
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor("#2B6CB0"),
        spaceAfter=6,
        spaceBefore=10
    )

    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#2D3748")
    )
    
    bold_style = ParagraphStyle(
        'BoldStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor("#1A202C")
    )

    story = []

    # Title
    story.append(Paragraph("B2P E-Commerce & Supply Chain SCM", title_style))
    story.append(Paragraph("INVOICE RECEIPT", ParagraphStyle('Sub', parent=title_style, fontSize=12, spaceAfter=20)))
    story.append(Spacer(1, 10))

    # Basic Info Table (Invoice Metadata)
    meta_data = [
        [Paragraph("<b>Order ID:</b>", normal_style), Paragraph(f"#{order.id}", normal_style),
         Paragraph("<b>Date:</b>", normal_style), Paragraph(timezone.localtime(order.order_confirmed_at).strftime('%Y-%m-%d %H:%M:%S'), normal_style)],
        [Paragraph("<b>Status:</b>", normal_style), Paragraph(order.get_status_display().upper(), bold_style),
         Paragraph("<b>Payment Method:</b>", normal_style), Paragraph(order.payment_method, normal_style)]
    ]
    t_meta = Table(meta_data, colWidths=[100, 160, 100, 160])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F7FAFC")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 15))

    # Parties Section (Buyer & Vendor)
    story.append(Paragraph("PARTIES INVOLVED", heading_style))
    parties_data = [
        [Paragraph("<b>BUYER (CLIENT)</b>", bold_style), Paragraph("<b>VENDOR (SUPPLIER)</b>", bold_style)],
        [
            Paragraph(f"Name: {order.buyer.username}<br/>Email: {order.buyer.email}", normal_style),
            Paragraph(f"Name: {order.product.vendor.username}<br/>Email: {order.product.vendor.email}", normal_style)
        ]
    ]
    t_parties = Table(parties_data, colWidths=[260, 260])
    t_parties.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EDF2F7")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 8),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#CBD5E0")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(t_parties)
    story.append(Spacer(1, 15))

    # Product details
    story.append(Paragraph("ORDER ITEMS & FINANCIALS", heading_style))
    financials_data = [
        [Paragraph("<b>Product / Item</b>", bold_style), Paragraph("<b>Unit Price (INR)</b>", bold_style), Paragraph("<b>Qty</b>", bold_style), Paragraph("<b>Total Price (INR)</b>", bold_style)],
        [
            Paragraph(order.product.name, normal_style), 
            Paragraph(f"INR {order.product.discounted_price}", normal_style),
            Paragraph(str(order.quantity), normal_style),
            Paragraph(f"INR {order.total_price}", bold_style)
        ]
    ]
    t_financials = Table(financials_data, colWidths=[220, 100, 60, 140])
    t_financials.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E2E8F0")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor("#CBD5E0")),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
    ]))
    story.append(t_financials)
    story.append(Spacer(1, 15))

    # Supply Chain Tracking History
    story.append(Paragraph("SUPPLY CHAIN TRACKING HISTORY LOGS", heading_style))
    
    def format_ts(ts):
        return timezone.localtime(ts).strftime('%Y-%m-%d %H:%M:%S') if ts else 'PENDING'

    tracking_data = [
        [Paragraph("<b>Milestone / Step</b>", bold_style), Paragraph("<b>Status</b>", bold_style), Paragraph("<b>Timestamp (IST)</b>", bold_style)],
        [Paragraph("1. Order Placement", normal_style), Paragraph("Confirmed", normal_style), Paragraph(format_ts(order.order_confirmed_at), normal_style)],
        [Paragraph("2. Logistics Dispatch", normal_style), Paragraph("Shipped", normal_style), Paragraph(format_ts(order.shipped_at), normal_style)],
        [Paragraph("3. Out For Delivery", normal_style), Paragraph("Out for Delivery", normal_style), Paragraph(format_ts(order.out_for_delivery_at), normal_style)],
        [Paragraph("4. End Customer Handover", normal_style), Paragraph("Delivered", normal_style), Paragraph(format_ts(order.delivered_at), normal_style)]
    ]
    t_tracking = Table(tracking_data, colWidths=[180, 140, 200])
    t_tracking.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EDF2F7")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E0")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(t_tracking)
    story.append(Spacer(1, 30))

    # Footer
    story.append(Paragraph("Thank you for using B2P Supply Chain SCM Portal.", ParagraphStyle('Footer', parent=normal_style, alignment=1, fontSize=8, textColor=colors.gray)))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def admin_approve_vendor(request, vendor_id):
    if not request.user.is_superuser and request.user.role != 'admin':
        messages.error(request, "Permission denied. Exclusive to Admin.")
        return redirect('home')
        
    target_user = get_object_or_404(UserProfile, id=vendor_id)
    if target_user == request.user:
        messages.error(request, "You cannot approve yourself.")
        return redirect('dashboard')
        
    if target_user.role == 'admin' or target_user.is_superuser:
        messages.error(request, "You cannot approve another administrator account.")
        return redirect('dashboard')
        
    if target_user.role != 'vendor':
        messages.error(request, "Only vendor accounts can be approved.")
        return redirect('dashboard')
        
    target_user.is_approved = True
    target_user.save()
    
    # Propagate approval to associated logistics employees
    UserProfile.objects.filter(created_by=target_user).update(is_approved=True)
    
    messages.success(request, f"Vendor '{target_user.username}' and their logistics employees approved successfully.")
    return redirect('/dashboard/?tab=vendors')

@login_required
def admin_disapprove_vendor(request, vendor_id):
    if not request.user.is_superuser and request.user.role != 'admin':
        messages.error(request, "Permission denied. Exclusive to Admin.")
        return redirect('home')
        
    target_user = get_object_or_404(UserProfile, id=vendor_id)
    if target_user == request.user:
        messages.error(request, "You cannot disapprove yourself.")
        return redirect('dashboard')
        
    if target_user.role == 'admin' or target_user.is_superuser:
        messages.error(request, "You cannot disapprove another administrator account.")
        return redirect('dashboard')
        
    if target_user.role != 'vendor':
        messages.error(request, "Only vendor accounts can be disapproved.")
        return redirect('dashboard')
        
    target_user.is_approved = False
    target_user.save()
    
    # Propagate disapproval: disapproving a vendor automatically marks all associated logistics accounts as disapproved (is_approved = False)
    UserProfile.objects.filter(created_by=target_user).update(is_approved=False)
    
    messages.warning(request, f"Vendor '{target_user.username}' disapproved. All associated logistics staff have been disapproved.")
    return redirect('/dashboard/?tab=vendors')

@login_required
def admin_create_admin(request):
    if not request.user.is_superuser and request.user.role != 'admin':
        messages.error(request, "Permission denied. Exclusive to Admin.")
        return redirect('home')
        
    if request.method == 'POST':
        form = AdminCreationForm(request.POST, request.FILES)
        if form.is_valid():
            admin_user = form.save()
            messages.success(request, f"Admin account '{admin_user.username}' created successfully!")
            return redirect('/dashboard/?tab=create_admin')
        else:
            messages.error(request, "Error creating admin account. Please check details.")
            vendors = UserProfile.objects.filter(role='vendor').order_by('-id')
            employees = UserProfile.objects.filter(role__in=['shipper', 'out_for_delivery', 'delivered']).order_by('-id')
            all_vendors = UserProfile.objects.filter(role='vendor').order_by('username')
            return render(request, 'core/admin_dashboard.html', {
                'vendors': vendors,
                'employees': employees,
                'all_vendors': all_vendors,
                'admin_form': form,
                'active_tab': 'create_admin'
            })
    return redirect('dashboard')

@login_required
def export_audit_report_excel(request):
    if not request.user.is_superuser and request.user.role != 'admin':
        messages.error(request, "Permission denied. Exclusive to Admin.")
        return redirect('home')
        
    # Get the same filters as the dashboard so the report matches the filtered view
    search_employees = request.GET.get('search_employees', '').strip()
    filter_vendor_id = request.GET.get('filter_vendor', '').strip()
    
    employees = UserProfile.objects.filter(role__in=['shipper', 'out_for_delivery', 'delivered']).order_by('-id')
    
    if search_employees:
        employees = employees.filter(
            Q(username__icontains=search_employees) |
            Q(email__icontains=search_employees) |
            Q(role__icontains=search_employees)
        )
        
    if filter_vendor_id:
        employees = employees.filter(created_by_id=filter_vendor_id)
        
    # Create an in-memory output file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logistics Staff Audit Register"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Design Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_title = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_header = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    
    border_thin = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin', color="CBD5E0"),
        right=openpyxl.styles.borders.Side(style='thin', color="CBD5E0"),
        top=openpyxl.styles.borders.Side(style='thin', color="CBD5E0"),
        bottom=openpyxl.styles.borders.Side(style='thin', color="CBD5E0")
    )
    
    # Title Block
    ws.merge_cells("A1:F1")
    ws["A1"] = "LOGISTICS STAFF AUDITING REGISTER"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Date of Export Info Block
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Export Date: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')} (IST)"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        "Creator Vendor", 
        "Employee Username", 
        "Email ID", 
        "Logistics Role", 
        "Status", 
        "Plain Password (Debugging)"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
    ws.row_dimensions[3].height = 25
    
    # Populate Data
    for row_idx, emp in enumerate(employees, 4):
        creator = emp.created_by.username if emp.created_by else "None (Admin Created)"
        role_disp = emp.get_role_display()
        
        # Determine status display
        if emp.created_by and not emp.created_by.is_approved:
            status = "Waiting for approval (Vendor Disapproved)"
        elif not emp.is_approved:
            status = "Waiting for approval"
        else:
            status = "Approved"
            
        password = emp.raw_password_view or "********"
        
        row_data = [creator, emp.username, emp.email, role_disp, status, password]
        
        is_even = (row_idx % 2 == 0)
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font_body
            cell.border = border_thin
            
            # Formatting specifics
            if col_idx in [1, 2]:
                cell.font = font_bold
            if col_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="center")
            if is_even:
                cell.fill = fill_zebra
                
        ws.row_dimensions[row_idx].height = 20
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            # Avoid using merged cell width calculation which can distort the column layout
            if cell.coordinate in ["A1", "B1", "C1", "D1", "E1", "F1", "A2", "B2", "C2", "D2", "E2", "F2"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Prepare HTTP Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Logistics_Staff_Audit_Report.xlsx"'
    
    # Save output to buffer and write to response
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    buffer.close()
    
    return response
